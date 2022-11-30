# Author: tbry3467 on GitHub
# re-coded original Java house cleaning project in Python w/ added functionality
# See the readme for objectives and requirements 
# todo: add a catch if all jobs are full, put everyone else in nojob

from os import name
import random
from datetime import date
from openpyxl import load_workbook


class Person:

    """ Used to regulate which members have been assigned what jobs
        each obj contains name and jobs1-4, represented as an int
        pass 0 for jobs1-4 when creating an obj, set later when assigned jobs
        ex: '5' would represent 'trash' has been assigned
    """

    def __init__(self, name, job1, job2, job3, job4): 
        # note: single underscore signifies to readers that it's a private attribute
        self.name = name
        self.job1 = job1
        self.job2 = job2
        self.job3 = job3
        self.job4 = job4


# lists used to track num of workers per job 
first_bathroom = []         # capacity 2
second_bathroom = []        # capacity 3
kitchen = []                # capacity 4
trash = []                  # capacity 1
attic = []                  # capacity 3
eagle_nest = []             # capacity 2
chapter_room = []           # capacity 1
odd_jobs = []               # capacity 5
first_floor_sweep = []      # capacity 1
first_floor_swiffer = []    # capacity 1
second_floor_sweep = []     # capacity 1
second_floor_swiffer = []   # capacity 1
no_job = []                 # no capacity, manual placement via House Manager

# array of all jobs, used for random selection
all_jobs = [first_bathroom, second_bathroom, kitchen, trash, attic,
eagle_nest, chapter_room, odd_jobs, first_floor_sweep, 
first_floor_swiffer, second_floor_sweep, second_floor_swiffer]

# reading members and adding to list
member_objects = []
in_file = "akl_member_names.txt" # "akl_member_names.txt"

with open(in_file) as f:
    content = f.read().splitlines() 

for line in content: # pull names, create member objects w default values and add to list
    line = Person(line,0,0,0,0)
    member_objects.append(line)

##### main loop #####

# todo: add a catch where if all jobs are full, put everyone else in no job

for x in range (1,5): # loops 4 times for 4 job assignments per person

    for ppl in member_objects:
        # assignment phase
        control = 0
        job_nums = [1,2,3,4,5,6,7,8,9,10,11,12] # jobs mapped to num, remove num when job is full
        capacity = [2,3,4,1,3,2,1,5,1,1,1,1] # num of ppl allowed in each job, in order

        while control == 0:
            random_job = int(random.random() * len(job_nums)) # num btwn 1 and len of job_num, adjusts when nums get removed

            for job in job_nums:
                if job == random_job:

                    # check if job is at max capacity
                    if len(all_jobs[random_job]) == capacity[random_job]: # if len(job) == capacity for said job:
                        job_nums.remove(job) # removes job num b/c full
                        del capacity[job-1] # removes corresponding capacity for that job
                        continue

                    # if the job is not at max capacity
                    else: 
                        if ppl.job1 == job or ppl.job2 == job or ppl.job3 == job or ppl.job4 == job: # prevents duplicate jobs w/in 1 month
                            continue

                        if x == 1: ppl.job1 = random_job # record job int in person's properties
                        elif x == 2: ppl.job2 = random_job
                        elif x == 3: ppl.job3 = random_job
                        else: ppl.job4 = random_job

                        (all_jobs[random_job]).append(ppl)
                        control=1

    # clear names from jobs to restart for the next week
    for jobs in all_jobs:
        jobs.clear

# test to make sure script works
for ppl in member_objects:
    print(ppl.name, ppl.job1, ppl.job2, ppl.job3, ppl.job4)

# writing phase
wb = load_workbook("test_output.xlsx") # change to output when done
sheet = wb.worksheets[0]

# clear cells from previous month - need to verify this works
for c in range (2,6):
    for r in range (2,59):
        sheet.cell(column=c, row=r).value == None

# week 1 writing
week_incrementer = 0 # used to regulate cells being filled

for col in range(2,6):
    for ppl in member_objects:
        if col == 2:
            job = ppl.job1
        elif col == 3:
            job = ppl.job2
        elif col == 4:
            job = ppl.job3
        else:
            job = ppl.job4

        # this check sequence ensures we can fit 1-5 people per job onto sheet (won't overwrite any names)
        if sheet.cell(column=2, row=job+1+week_incrementer).value == None:
            sheet.cell(column=2, row=job+1+week_incrementer).value = ppl.name
        elif sheet.cell(column=3, row=job+1+week_incrementer).value == None:
            sheet.cell(column=3, row=job+1+week_incrementer).value = ppl.name
        elif sheet.cell(column=4, row=job+1+week_incrementer).value == None:
            sheet.cell(column=4, row=job+1+week_incrementer).value = ppl.name
        elif sheet.cell(column=5, row=job+1+week_incrementer).value == None:
            sheet.cell(column=5, row=job+1+week_incrementer).value = ppl.name
        else:
            sheet.cell(column=6, row=job+1+week_incrementer).value = ppl.name
    week_incrementer += 15


# set headers to proper date
# mondays = get_mondays()

wb.save("test_output.xlsx")